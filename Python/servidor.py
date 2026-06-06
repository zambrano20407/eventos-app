#!/usr/bin/env python3
import os, sys, json, base64, io, shutil, threading, webbrowser
from http.server import HTTPServer, SimpleHTTPRequestHandler
from datetime import datetime
from urllib.parse import urlparse, parse_qs

# ── Verificar dependencias ───────────────────────────────────
def verificar_deps():
    faltantes = []
    for pkg in ['firebase_admin', 'openpyxl', 'PIL']:
        try:
            __import__(pkg)
        except ImportError:
            nombre = {'firebase_admin':'firebase-admin','PIL':'Pillow'}.get(pkg,pkg)
            faltantes.append(nombre)
    if faltantes:
        print(f"\n❌ Faltan librerias. Ejecute:\n   py -m pip install {' '.join(faltantes)}\n")
        sys.exit(1)

verificar_deps()

import firebase_admin
from firebase_admin import credentials, firestore
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

# ══════════════════════════════════════════════════════════════
#  RUTAS — servidor.py esta en /eventos-app/Python/
#  La raiz del proyecto es un nivel arriba
# ══════════════════════════════════════════════════════════════
PYTHON_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR   = os.path.dirname(PYTHON_DIR)

RUTA_CREDENCIALES = os.path.join(PYTHON_DIR, 'serviceAccountKey.json')
RUTA_TEMPLATE     = os.path.join(PYTHON_DIR, 'Formato_PTFT38.xlsx')
CARPETA_REPORTES  = os.path.join(PYTHON_DIR, 'reportes')
PUERTO            = 8000

# ── Conectar Firebase ─────────────────────────────────────────
def conectar_firebase():
    if not firebase_admin._apps:
        if not os.path.exists(RUTA_CREDENCIALES):
            print(f"❌ No se encontro: {RUTA_CREDENCIALES}")
            sys.exit(1)
        cred = credentials.Certificate(RUTA_CREDENCIALES)
        firebase_admin.initialize_app(cred)
    return firestore.client()

# ── Insertar firma ────────────────────────────────────────────
def insertar_firma(ws, firma_b64, fila):
    if not firma_b64 or ',' not in firma_b64:
        return False
    try:
        b64_data = firma_b64.split(',')[1]
        img_bytes = base64.b64decode(b64_data)
        pil_img = PILImage.open(io.BytesIO(img_bytes)).convert('RGBA')
        fondo = PILImage.new('RGBA', pil_img.size, (255,255,255,255))
        fondo.paste(pil_img, mask=pil_img.split()[3])
        pil_img = fondo.convert('RGB')
        pil_img.thumbnail((200, 55), PILImage.LANCZOS)
        out_buf = io.BytesIO()
        pil_img.save(out_buf, format='PNG')
        out_buf.seek(0)
        xl_img = XLImage(out_buf)
        xl_img.anchor = f'O{fila}'
        ws.add_image(xl_img)
        return True
    except Exception as e:
        print(f"  ⚠️  Firma fila {fila}: {e}")
        return False

# ── Generar Excel ─────────────────────────────────────────────
def generar_excel(evento, registros):
    if not os.path.exists(RUTA_TEMPLATE):
        return None, f"No se encontro el template: {RUTA_TEMPLATE}"

    os.makedirs(CARPETA_REPORTES, exist_ok=True)
    nombre = (evento.get('nombre') or 'evento').replace(' ','_').replace('/','_')
    fecha  = datetime.now().strftime('%Y%m%d_%H%M')
    salida = os.path.join(CARPETA_REPORTES, f'PTFT38_{nombre}_{fecha}.xlsx')

    shutil.copy2(RUTA_TEMPLATE, salida)
    wb = openpyxl.load_workbook(salida)
    ws = wb.active

    ws['B9'].value  = f"Institucion que dicta el curso / formacion / capacitacion:  {evento.get('institucion','')}"
    ws['I9'].value  = f"Fecha:  {evento.get('fecha','')}"
    ws['O9'].value  = f"Jornada:  {evento.get('jornada','')}"
    ws['B10'].value = f"Nombre del curso / formacion / capacitacion:  {evento.get('nombre','')}"

    # Limpiar filas de datos antes de escribir (evita datos residuales de la plantilla)
    COLUMNAS_DATOS = [3, 5, 7, 8, 9, 10, 11, 12, 13, 15]
    for fila_limpia in range(13, 38):
        for col in COLUMNAS_DATOS:
            ws.cell(row=fila_limpia, column=col).value = None

    con_firma = 0
    for i, reg in enumerate(registros[:25]):
        fila  = 13 + i
        nivel = (reg.get('nivel') or '').lower()
        ws.cell(row=fila, column=3).value  = reg.get('cedula','')
        ws.cell(row=fila, column=5).value  = reg.get('nombre','')
        ws.cell(row=fila, column=7).value  = reg.get('dependencia','')
        ws.cell(row=fila, column=8).value  = (reg.get('sexo') or 'M')[0].upper()
        ws.cell(row=fila, column=9 ).value = 'X' if nivel == 'directivo'   else ''
        ws.cell(row=fila, column=10).value = 'X' if nivel == 'asesor'      else ''
        ws.cell(row=fila, column=11).value = 'X' if nivel == 'profesional' else ''
        ws.cell(row=fila, column=12).value = 'X' if nivel == 'tecnico'     else ''
        ws.cell(row=fila, column=13).value = 'X' if nivel == 'asistencial' else ''
        if insertar_firma(ws, reg.get('firma',''), fila):
            con_firma += 1

    wb.save(salida)
    print(f"  ✅ Excel generado: {salida} ({con_firma} firmas)")
    return salida, None

# ── Firebase helpers ──────────────────────────────────────────
def obtener_registros(db, evento_id):
    docs = db.collection(f'eventos/{evento_id}/registros').order_by('creadoEn').stream()
    return [{'id': d.id, **d.to_dict()} for d in docs]

def obtener_evento(db, evento_id):
    doc = db.collection('eventos').document(evento_id).get()
    if doc.exists:
        return {'id': doc.id, **doc.to_dict()}
    return None

# ══════════════════════════════════════════════════════════════
#  SERVIDOR HTTP
# ══════════════════════════════════════════════════════════════
db = None

class Handler(SimpleHTTPRequestHandler):

    def log_message(self, format, *args):
        try:
            if args and isinstance(args[0], str) and '/api/' in args[0]:
                print(f"  [{datetime.now().strftime('%H:%M:%S')}] {args[0]}")
        except Exception:
            pass

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

    def do_GET(self):
        parsed = urlparse(self.path)

        # ── API exportar evento ───────────────────────────────
        if parsed.path == '/api/exportar':
            params = parse_qs(parsed.query)
            ev_id  = params.get('eventoId', [None])[0]
            if not ev_id:
                self._error(400, 'Falta eventoId')
                return
            print(f"\n📥 Exportando evento: {ev_id}")
            try:
                evento    = obtener_evento(db, ev_id)
                if not evento:
                    self._error(404, 'Evento no encontrado')
                    return
                registros = obtener_registros(db, ev_id)
                if not registros:
                    self._error(404, 'El evento no tiene registros')
                    return
                ruta, error = generar_excel(evento, registros)
                if error:
                    self._error(500, error)
                    return
                nombre_archivo = f'PTFT38_{(evento.get("nombre") or "evento").replace(" ","_")}.xlsx'
                with open(ruta, 'rb') as f:
                    contenido = f.read()
                self.send_response(200)
                self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header('Content-Disposition', f'attachment; filename="{nombre_archivo}"')
                self.send_header('Content-Length', str(len(contenido)))
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(contenido)
            except Exception as e:
                print(f"  ❌ Error: {e}")
                self._error(500, str(e))
            return

        # ── API exportar todos ────────────────────────────────
        elif parsed.path == '/api/exportar-todo':
            print(f"\n📥 Exportando todos los eventos")
            try:
                todos = []
                evento_gen = {
                    'nombre':      'Todos los Eventos',
                    'fecha':       datetime.now().strftime('%d/%m/%Y'),
                    'institucion': 'Registraduria Nacional del Estado Civil',
                    'jornada':     ''
                }
                for doc in db.collection('eventos').stream():
                    ev   = {'id': doc.id, **doc.to_dict()}
                    regs = obtener_registros(db, doc.id)
                    for r in regs:
                        r['eventoNombre'] = ev.get('nombre','')
                        r['eventoFecha']  = ev.get('fecha','')
                        todos.append(r)
                if not todos:
                    self._error(404, 'No hay registros')
                    return
                ruta, error = generar_excel(evento_gen, todos)
                if error:
                    self._error(500, error)
                    return
                with open(ruta, 'rb') as f:
                    contenido = f.read()
                self.send_response(200)
                self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header('Content-Disposition', 'attachment; filename="PTFT38_Todos_los_Eventos.xlsx"')
                self.send_header('Content-Length', str(len(contenido)))
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(contenido)
            except Exception as e:
                print(f"  ❌ Error: {e}")
                self._error(500, str(e))
            return

        # ── Archivos estaticos ────────────────────────────────
        else:
            super().do_GET()

    def _error(self, code, msg):
        body = json.dumps({'error': msg}).encode()
        self.send_response(code)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Content-Length', str(len(body)))
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(body)


# ══════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════
def main():
    global db

    print("\n╔══════════════════════════════════════════════════════╗")
    print("║   SISTEMA PTFT38 — Registraduria Nacional Caqueta   ║")
    print("╚══════════════════════════════════════════════════════╝\n")

    print("🔌 Conectando a Firebase...")
    db = conectar_firebase()
    print("   ✅ Firebase conectado\n")

    if not os.path.exists(RUTA_TEMPLATE):
        print(f"⚠️  No se encontro {RUTA_TEMPLATE}")
    else:
        print(f"   ✅ Template encontrado\n")

    # Cambiar al directorio raiz para servir archivos estaticos
    os.chdir(BASE_DIR)

    servidor = HTTPServer(('localhost', PUERTO), Handler)
    url      = f'http://localhost:{PUERTO}/vistas/admin.html'

    print(f"🚀 Servidor iniciado en: http://localhost:{PUERTO}")
    print(f"📋 Panel Admin:          {url}")
    print(f"\n   Presione Ctrl+C para detener el servidor")
    print("─" * 54 + "\n")

    threading.Timer(1.0, lambda: webbrowser.open(url)).start()

    try:
        servidor.serve_forever()
    except KeyboardInterrupt:
        print("\n\n👋 Servidor detenido.")
        servidor.server_close()


if __name__ == '__main__':
    main()
