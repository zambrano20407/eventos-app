#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════╗
║  GENERADOR DE REPORTE PTFT38 — Registraduría Caquetá        ║
║  Conecta a Firebase, descarga los registros de un evento    ║
║  y genera el formato institucional Excel con firmas         ║
╚══════════════════════════════════════════════════════════════╝

INSTALACIÓN (solo la primera vez):
  pip install firebase-admin openpyxl Pillow

USO:
  python3 generar_reporte.py
  → Le preguntará qué evento exportar y generará el Excel
"""

import os, sys, base64, io, shutil
from datetime import datetime

# ── Verificar dependencias ───────────────────────────────────
def verificar_deps():
    faltantes = []
    for pkg in ['firebase_admin', 'openpyxl', 'PIL']:
        try:
            __import__(pkg)
        except ImportError:
            nombre = {'firebase_admin':'firebase-admin', 'PIL':'Pillow'}.get(pkg, pkg)
            faltantes.append(nombre)
    if faltantes:
        print(f"\n❌ Faltan librerías. Ejecute:\n   pip install {' '.join(faltantes)}\n")
        sys.exit(1)

verificar_deps()

import firebase_admin
from firebase_admin import credentials, firestore
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

# ══════════════════════════════════════════════════════════════
#  CONFIGURACIÓN — Edite estas rutas
# ══════════════════════════════════════════════════════════════
# Ruta al archivo de credenciales de Firebase
# Descárguelo en: Firebase Console → Configuración → Cuentas de servicio → Generar clave privada
RUTA_CREDENCIALES = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'serviceAccountKey.json')

# Ruta al formato institucional Excel (el que subió)
RUTA_TEMPLATE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Formato_PTFT38.xlsx')

# Carpeta donde se guardarán los reportes generados
CARPETA_REPORTES = 'reportes'
# ══════════════════════════════════════════════════════════════


def conectar_firebase():
    """Conecta a Firebase usando las credenciales."""
    if not os.path.exists(RUTA_CREDENCIALES):
        print(f"""
❌ No se encontró el archivo de credenciales: {RUTA_CREDENCIALES}

Para obtenerlo:
  1. Vaya a Firebase Console → su proyecto
  2. ⚙️ Configuración → Cuentas de servicio
  3. Haga clic en "Generar nueva clave privada"
  4. Guarde el archivo como: {RUTA_CREDENCIALES}
  5. Colóquelo en la misma carpeta que este script
""")
        sys.exit(1)

    if not firebase_admin._apps:
        cred = credentials.Certificate(RUTA_CREDENCIALES)
        firebase_admin.initialize_app(cred)

    return firestore.client()


def listar_eventos(db):
    """Lista todos los eventos disponibles."""
    eventos = []
    docs = db.collection('eventos').order_by('creadoEn', direction=firestore.Query.DESCENDING).stream()
    for doc in docs:
        ev = doc.to_dict()
        ev['id'] = doc.id
        # Contar registros
        regs = db.collection(f'eventos/{doc.id}/registros').stream()
        ev['total_registros'] = sum(1 for _ in regs)
        eventos.append(ev)
    return eventos


def descargar_registros(db, evento_id):
    """Descarga todos los registros de un evento."""
    registros = []
    docs = db.collection(f'eventos/{evento_id}/registros').order_by('creadoEn').stream()
    for doc in docs:
        reg = doc.to_dict()
        reg['id'] = doc.id
        registros.append(reg)
    return registros


def insertar_firma(ws, firma_b64, fila):
    """Inserta la imagen de firma en la celda O de la fila indicada."""
    if not firma_b64 or ',' not in firma_b64:
        return False
    try:
        b64_data  = firma_b64.split(',')[1]
        img_bytes = base64.b64decode(b64_data)

        pil_img = PILImage.open(io.BytesIO(img_bytes)).convert('RGBA')

        # Quitar transparencia (fondo blanco)
        fondo = PILImage.new('RGBA', pil_img.size, (255, 255, 255, 255))
        fondo.paste(pil_img, mask=pil_img.split()[3])
        pil_img = fondo.convert('RGB')

        # Redimensionar para que quepa en la celda (~200x55px)
        pil_img.thumbnail((200, 55), PILImage.LANCZOS)

        out_buf = io.BytesIO()
        pil_img.save(out_buf, format='PNG')
        out_buf.seek(0)

        xl_img        = XLImage(out_buf)
        xl_img.anchor = f'O{fila}'
        ws.add_image(xl_img)
        return True

    except Exception as e:
        print(f"    ⚠️  Error insertando firma fila {fila}: {e}")
        return False


def generar_excel(evento, registros):
    """Llena el formato institucional con los datos y firmas."""

    if not os.path.exists(RUTA_TEMPLATE):
        print(f"❌ No se encontró el template: {RUTA_TEMPLATE}")
        sys.exit(1)

    # Crear carpeta de reportes
    os.makedirs(CARPETA_REPORTES, exist_ok=True)

    # Nombre del archivo de salida
    nombre_evento = (evento.get('nombre') or 'evento').replace(' ','_').replace('/','_')
    fecha_hoy     = datetime.now().strftime('%Y%m%d_%H%M')
    ruta_salida   = os.path.join(CARPETA_REPORTES, f'PTFT38_{nombre_evento}_{fecha_hoy}.xlsx')

    # Copiar template sin tocarlo
    shutil.copy2(RUTA_TEMPLATE, ruta_salida)
    wb = openpyxl.load_workbook(ruta_salida)
    ws = wb.active

    # ── Datos del evento ──────────────────────────────────────
    ws['B9'].value  = f"Institución que dicta el curso / formación / capacitación:  {evento.get('institucion','')}"
    ws['I9'].value  = f"Fecha:  {evento.get('fecha','')}"
    ws['O9'].value  = f"Jornada:  {evento.get('jornada','')}"
    ws['B10'].value = f"Nombre del curso / formación / capacitación:  {evento.get('nombre','')}"

    # ── Registros (máx 25 por hoja) ───────────────────────────
    total     = len(registros)
    llenados  = 0
    con_firma = 0

    for i, reg in enumerate(registros[:25]):
        fila  = 13 + i
        nivel = (reg.get('nivel') or '').lower()

        ws.cell(row=fila, column=3).value  = reg.get('cedula', '')
        ws.cell(row=fila, column=5).value  = reg.get('nombre', '')
        ws.cell(row=fila, column=7).value  = reg.get('dependencia', '')
        ws.cell(row=fila, column=8).value  = (reg.get('sexo') or 'M')[0].upper()

        ws.cell(row=fila, column=9 ).value = 'X' if nivel == 'directivo'   else ''
        ws.cell(row=fila, column=10).value = 'X' if nivel == 'asesor'      else ''
        ws.cell(row=fila, column=11).value = 'X' if nivel == 'profesional' else ''
        ws.cell(row=fila, column=12).value = 'X' if nivel == 'técnico'     else ''
        ws.cell(row=fila, column=13).value = 'X' if nivel == 'asistencial' else ''

        # Firma como imagen
        if insertar_firma(ws, reg.get('firma', ''), fila):
            con_firma += 1

        llenados += 1

    wb.save(ruta_salida)

    # ── Resumen ────────────────────────────────────────────────
    print(f"""
✅ Reporte generado exitosamente:
   📄 Archivo:    {ruta_salida}
   👥 Registros:  {llenados} de {total}
   ✍️  Con firma:  {con_firma}
   ⚠️  Sin firma:  {llenados - con_firma}
""")
    if total > 25:
        print(f"⚠️  Nota: El formato tiene máximo 25 filas. Quedaron {total-25} registros sin incluir.")

    return ruta_salida


def main():
    print("""
╔══════════════════════════════════════════════════════════════╗
║        GENERADOR DE REPORTE PTFT38 — Registraduría          ║
╚══════════════════════════════════════════════════════════════╝
""")

    # Conectar a Firebase
    print("🔌 Conectando a Firebase...")
    db = conectar_firebase()
    print("   ✅ Conectado\n")

    # Listar eventos
    print("📋 Cargando eventos...")
    eventos = listar_eventos(db)

    if not eventos:
        print("❌ No hay eventos registrados en Firebase.")
        sys.exit(0)

    print(f"\n{'#':<4} {'Nombre':<35} {'Fecha':<12} {'Registros':<10} {'Jornada'}")
    print("─" * 75)
    for idx, ev in enumerate(eventos):
        nombre = (ev.get('nombre') or '')[:34]
        fecha  = ev.get('fecha', '—')
        total  = ev.get('total_registros', 0)
        jorn   = ev.get('jornada', '—')
        print(f"{idx+1:<4} {nombre:<35} {fecha:<12} {total:<10} {jorn}")

    # Seleccionar evento
    print()
    while True:
        try:
            sel = input("Ingrese el número del evento a exportar (o 0 para todos): ").strip()
            sel = int(sel)
            if 0 <= sel <= len(eventos):
                break
            print(f"   Por favor ingrese un número entre 0 y {len(eventos)}")
        except ValueError:
            print("   Por favor ingrese un número válido")

    # Generar reporte(s)
    if sel == 0:
        # Exportar todos uno por uno
        for ev in eventos:
            if ev['total_registros'] == 0:
                print(f"   ⏭️  Saltando '{ev.get('nombre')}' (sin registros)")
                continue
            print(f"\n📥 Exportando: {ev.get('nombre')}...")
            regs = descargar_registros(db, ev['id'])
            generar_excel(ev, regs)
    else:
        ev   = eventos[sel - 1]
        print(f"\n📥 Descargando registros de: {ev.get('nombre')}...")
        regs = descargar_registros(db, ev['id'])
        if not regs:
            print("❌ Este evento no tiene registros.")
            sys.exit(0)
        generar_excel(ev, regs)

    print("🎉 ¡Listo! Revise la carpeta 'reportes'")


if __name__ == '__main__':
    main()